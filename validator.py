import openpyxl

import re

import os

from datetime import datetime

import openpyxl.utils

from openpyxl.utils import get_column_letter

# ANSI escape codes for color output

RED = "\033[91m"  # Critical (Red)

YELLOW = "\033[93m"  # Warning (Yellow)

BLUE = "\033[94m"  # Info (Blue)

GREEN = "\033[92m"  # Success (Green)

RESET = "\033[0m"  # Reset color to default

# Dictionary to store validation errors

errors = {

   "Critical": [],

   "Warning": [],

   "Info": []

}

def validate_excel(file_path):
    """ Validates the Excel file based on provided rules. """
    try:
        wb = openpyxl.load_workbook(file_path)

        ### 🔵 Empty Cell Validation (`excel_rule_1`) ###

        """"
        excel_rule_1: Non-Empty Cells with Specific Exceptions

        Rule Definition:
            All data cells within the specified Excel sheets must contain a value. 
            The first row of each sheet is designated as the header row and is therefore excluded from this validation.
            Data validation will commence from the second row onwards.
            
        Exceptions and Warnings:

            The following columns and cells are exempt from the non-empty cell requirement, and any deviations from the specified conditions 
            should be flagged as warnings:
                1. "swc_info" Sheet:
                    * Columns D and I: These columns may contain empty cells.
                    * Column M: The content of this column is contingent upon the value in column L, as detailed below:
                    * For column L values of 'AsynchronousServerCallReturnsEvent', 'BackgroundEvent', 'InitEvent', 'InternalTriggerOccurredEvent',
                        'OsTaskExecutionEvent', 'SwcModeManagerErrorEvent' and 'TransformerHardErrorEvent', column M must be empty.
                    * For column L values of 'DataReceivedEvent', 'DataReceiveErrorEvent', 'DataSendCompletedEvent', 'DataWriteCompletedEvent',
                        'ExternalTriggerOccurredEvent', 'ModeSwitchedAckEvent', 'OperationInvokedEvent', and 'SwcModeSwitchEvent', column M must
                        contain a port name. This port name must correspond to a value found in column C of the "ports" sheet.
                            * Furthermore, the corresponding row in the "ports" sheet must satisfy the following criteria:
                            * For 'DataReceivedEvent' and 'DataReceiveErrorEvent' the corresponding "B" column value must be "ReceiverPort" and
                                "D" column value must be either "SenderReceiverInterface" or "NvDataInterface".
                            * For 'DataSendCompletedEvent' and 'DataWriteCompletedEvent' the corresponding "B" column value must be "SenderPort"
                                and "D" column value must be either "SenderReceiverInterface" or "NvDataInterface".
                            * For 'ExternalTriggerOccurredEvent' the corresponding "B" column value must be "ReceiverPort" and "D" column value 
                                must be "TriggerInterface".
                            * For 'ModeSwitchedAckEvent' and 'OperationInvokedEvent' the corresponding "B" column value must be "SenderPort" and
                                "D" column value must be "ModeSwitchInterface" and "ClientServerInterface" respectively.
                            * For 'SwcModeSwitchEvent' the corresponding "B" column value must be "ReceiverPort" and "D" column value must be
                                "ModeSwitchInterface".
                            * Example: If column M contains 'rport1', and 'rport1' is located in cell C7 of the "ports" sheet, then cells B7 and D7 of
                                the "ports" sheet must contain the respective corresponding values as described above.
                    * For column L value of 'TimingEvent' column M must contain a numeric time value (e.g., 1.0, 0.87).
                2. "ib_data" Sheet:
                    * Column E: This column may contain empty cells.
                    * Column M: If corresponding column B value contains either "PerInstanceMemory" or "ArTypedPerInstanceMemory", then column M
                                must be empty.
                3. "ports" Sheet:
                    * Columns J, K, and L: These columns may contain empty cells.
                4. "adt_primitive" Sheet:
                    * If column E contains the value 'IDENTICAL', then corresponding columns F and G values must be empty.
                5. "idt" Sheet:
                    * If column B contains the value 'PRIMITIVE', then corresponding column D value must be empty.

        """
        # Define sheets and exception columns
        empty_check_sheets = {
        "swc_info": ["D", "I", "M"],
        "ib_data": ["E"],
        "ports": ["J", "K", "L"],
        "adt_primitive": ["E", "F", "G"],
        "idt": ["B", "D"]
        }
        column_limits = {
        "swc_info": "M",
        "ib_data": "I",
        "ports": "L",
        "adt_primitive": "M",
        "adt_composite": "F",
        "idt": "E"
        }
        # Load sheets
        swc_info = wb["swc_info"] if "swc_info" in wb.sheetnames else None
        ports = wb["ports"] if "ports" in wb.sheetnames else None
        ib_data = wb["ib_data"] if "ib_data" in wb.sheetnames else None
        adt_primitive = wb["adt_primitive"] if "adt_primitive" in wb.sheetnames else None
        idt = wb["idt"] if "idt" in wb.sheetnames else None

        # Function to get merged cell mappings
        def get_merged_cells(sheet):
            merged_ranges = {}
            for merged_range in sheet.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                first_cell = f"{get_column_letter(min_col)}{min_row}"
                merged_cells = [
                    f"{get_column_letter(col)}{row}"
                    for col in range(min_col, max_col + 1)
                    for row in range(min_row, max_row + 1)
                    if f"{get_column_letter(col)}{row}" != first_cell
                ]
                merged_ranges[first_cell] = merged_cells
            return merged_ranges
        # 🔹 1️⃣ General non-empty validation
        for sheet_name, exception_columns in empty_check_sheets.items():
            sheet = wb[sheet_name] if sheet_name in wb.sheetnames else None
            if sheet:
                merged_ranges = get_merged_cells(sheet)
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    for col_idx, cell in enumerate(row):
                        column_letter = get_column_letter(col_idx + 1)
                        cell_ref = f"{column_letter}{row_idx}"
                        # Skip merged cells (except first cell)
                        if any(cell_ref in merged_cells for merged_cells in merged_ranges.values()):
                            errors["Info"].append(f"[{sheet_name}] Merged cell {cell_ref} is expected to be empty")
                            continue  
                        # Skip exception columns
                        if column_letter in exception_columns:
                            continue  
                        # Check for missing value
                        if cell.value in [None, ""]:
                            errors["Critical"].append(f"[{sheet_name}] Missing value at {cell_ref}")
        # 🔹 2️⃣ Get ports mapping
        def get_ports_mapping():
            if not ports:
                return {}
            ports_map = {}
            for row in ports.iter_rows(min_row=2, values_only=True):
                if len(row) > 3:
                    port_name, b_value, d_value = row[2], row[1], row[3]  # Columns C, B, D
                    if port_name:
                        ports_map[port_name] = (b_value, d_value)
            return ports_map
        ports_mapping = get_ports_mapping()
        # 🔹 3️⃣ Column M validation rules for swc_info
        m_validation_rules = {
        "AsynchronousServerCallReturnsEvent": lambda m: m is None,
        "BackgroundEvent": lambda m: m is None,
        "DataReceivedEvent": lambda m: ports_mapping.get(m, (None, None)) in [ ("ReceiverPort", "SenderReceiverInterface"), ("ReceiverPort", "NvDataInterface") ],
        "DataReceiveErrorEvent": lambda m: ports_mapping.get(m, (None, None)) in [ ("ReceiverPort", "SenderReceiverInterface"), ("ReceiverPort", "NvDataInterface") ],
        "DataSendCompletedEvent": lambda m: ports_mapping.get(m, (None, None)) in [ ("SenderPort", "SenderReceiverInterface"), ("SenderPort", "NvDataInterface")],
        "DataWriteCompletedEvent": lambda m: ports_mapping.get(m, (None, None)) in [ ("SenderPort", "SenderReceiverInterface"), ("SenderPort", "NvDataInterface") ],
        "ExternalTriggerOccurredEvent": lambda m: ports_mapping.get(m, (None, None)) == ("ReceiverPort", "TriggerInterface"),
        "InitEvent": lambda m: m is None,
        "InternalTriggerOccurredEvent": lambda m: m is None,
        "ModeSwitchedAckEvent": lambda m: ports_mapping.get(m, (None, None)) == ("SenderPort", "ModeSwitchInterface"),
        "OperationInvokedEvent": lambda m: ports_mapping.get(m, (None, None)) == ("SenderPort", "ClientServerInterface"),
        "OsTaskExecutionEvent": lambda m: m is None,
        "SwcModeManagerErrorEvent": lambda m: m is None,
        "SwcModeSwitchEvent": lambda m: ports_mapping.get(m, (None, None)) == ("ReceiverPort", "ModeSwitchInterface"),
        "TimingEvent": lambda m: isinstance(m, (int, float)),
        "TransformerHardErrorEvent": lambda m: m is None,
        }
        # Validate column M in swc_info
        if swc_info:
            for row_idx, row in enumerate(swc_info.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) > 12:
                    event_type = row[11]  # Column L
                    m_value = row[12]  # Column M
                    cell_ref = f"M{row_idx}"
                    if event_type in m_validation_rules and not m_validation_rules[event_type](m_value):
                        errors["Critical"].append(f"[swc_info] Invalid value at {cell_ref} for event type '{event_type}'")
            # 🔹 4️⃣ Validate column M in ib_data
        if ib_data:
            for row_idx, row in enumerate(ib_data.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) > 12:
                    col_b_value = row[1]  # Column B
                    col_m_value = row[12]  # Column M
                    cell_ref = f"M{row_idx}"
                    if col_b_value in ["PerInstanceMemory", "ArTypedPerInstanceMemory"] and col_m_value is not None:
                        errors["Critical"].append(f"[ib_data] Column M must be empty at {cell_ref} when Column B is '{col_b_value}'")

        # 🔹 5️⃣ Validate column E, F & G in adt_primitive
        if adt_primitive:
            merged_ranges = get_merged_cells(adt_primitive)  # Get merged ranges
            for row_idx, row in enumerate(adt_primitive.iter_rows(min_row=2), start=2):
                for col_idx in [4, 5, 6]:  # Columns E, F, G (0-based index: 4, 5, 6)
                    column_letter = get_column_letter(col_idx + 1)
                    cell_ref = f"{column_letter}{row_idx}"
                    cell = row[col_idx]
                    # Skip merged cells (except first cell)
                    if any(cell_ref in merged_cells for merged_cells in merged_ranges.values()):
                        errors["Info"].append(f"[adt_primitive] Merged cell {cell_ref} is expected to be empty")
                        continue  
                    # Validate based on rules
                    if column_letter == "E" and cell.value in [None, ""]:
                        errors["Critical"].append(f"[adt_primitive] Column E must not be empty at {cell_ref}")
                    elif column_letter in ["F", "G"] and row[4].value == "IDENTICAL" and cell.value not in [None, ""]:
                        errors["Critical"].append(f"[adt_primitive] Column {column_letter} must be empty at {cell_ref} when Column E is 'IDENTICAL'")
                                
        # 🔹 6️⃣Validate column B & D in idt
        if idt:
            merged_ranges = get_merged_cells(idt)  # Get merged ranges
            for row_idx, row in enumerate(idt.iter_rows(min_row=2), start=2):
                for col_idx in [1, 3]:  # Columns B, D (0-based index: 1, 3)
                    column_letter = get_column_letter(col_idx + 1)
                    cell_ref = f"{column_letter}{row_idx}"
                    cell = row[col_idx]
                    # Skip merged cells (except first cell)
                    if any(cell_ref in merged_cells for merged_cells in merged_ranges.values()):
                        errors["Info"].append(f"[idt] Merged cell {cell_ref} is expected to be empty")
                        continue  
                    # Validate based on rules
                    if column_letter == "B" and cell.value in [None, ""]:
                        errors["Critical"].append(f"[idt] Column B must not be empty at {cell_ref}")
                    elif column_letter == "D" and row[1].value == "PRIMITIVE" and cell.value not in [None, ""]:
                        errors["Critical"].append(f"[idt] Column D must be empty at {cell_ref} when Column B is 'PRIMITIVE'")

        ### 🟢 Naming Convention Rule ('excel_rule_2') ###

        """ 
        excel_rule_2 : Naming convention
            . this rule is applicable to following user given values             
                in 	"swc_info" column "C",  column "D", column "E", column "H", column "I", and column "K"
                in 	"ib_data" column "C"
                in 	"ports" column "C",  column "E", column "F", and column "G"
                in 	"adt_primitive" column "B",  column "D", column "G", and column "I"
                in 	"adt_composite" column "C" and  column "D"
                in 	"idt" column "C",  and column "D"
            . first row of every sheet will be the header so the data for validation should be consider from second row of each above mentioned excel sheets.
            . the rule is 'the name can have small and capital alphabetical letters and numbers from 0 to 9 and no special characters except _ '            
            . the name can start with only alphabetical which can be either capital or small letters 
        """

        naming_sheets = {
            "swc_info": ["C", "D", "E", "H", "I", "K"],
            "ib_data": ["C"],
            "ports": ["C", "E", "F", "G"],
            "adt_primitive": ["B", "D", "G", "I"],
            "adt_composite": ["C", "D"],
            "idt": ["C", "D"]
        }
        for sheet_name, columns in naming_sheets.items():
            if sheet_name not in wb.sheetnames:
                continue  # Skip if sheet doesn't exist
            sheet = wb[sheet_name]
            for col in columns:
                for row_idx, row in enumerate(sheet.iter_rows(
                        min_row=2, min_col=ord(col.upper()) - 64, max_col=ord(col.upper()) - 64, values_only=True), start=2):
                    name = row[0] if row else ""
                    cell_ref = f"{col}{row_idx}"
                    # Special handling for Column D in `adt_composite` and `idt`
                    if col == "D" and sheet_name in ["adt_composite", "idt"]:
                        if str(name).isdigit():  
                            # Log as info if it's purely numeric
                            errors["Info"].append(f"[{sheet_name}] Numeric value in naming column at {cell_ref}: {name}")
                            continue  # Skip further validation
                    # Apply normal naming convention check
                    if not re.match(r"^[A-Za-z][A-Za-z0-9_]*$", str(name)):
                        errors["Critical"].append(f"[{sheet_name}] Invalid name format at {cell_ref}: {name}")

        ### 🟡 Duplicate & Definition Consistency Rules ('excel_rule_3') ###
        duplicate_sheets = {
            "swc_info": ["H", "I", "K"],
            "ib_data": ["C"],
            "ports": ["C"],
            "adt_primitive": ["B", "G"],
            "adt_composite": ["C", "D"],  # D has special handling
            "idt": ["C", "D"]  # D has special handling
        }
        for sheet_name, columns in duplicate_sheets.items():
            sheet = wb[sheet_name]
            merged_ranges = get_merged_cells(sheet)  # Get merged cell mappings
            for col in columns:
                seen = set()
                for row_idx, row in enumerate(sheet.iter_rows(
                        min_row=2, min_col=ord(col.upper()) - 64, max_col=ord(col.upper()) - 64, values_only=True), start=2):
                    value = row[0]
                    cell_ref = f"{col}{row_idx}"
                    # Skip empty (None) values entirely
                    if value in [None, ""]:
                        continue  # Do not check empty values for duplication
                    # Handle merged cells (if part of a merged range, log as info and skip)
                    if any(cell_ref in merged_cells for merged_cells in merged_ranges.values()):
                        errors["Info"].append(f"[{sheet_name}] Merged cell {cell_ref} is expected to have the same value")
                        continue  # Skip checking duplicates for merged empty cells
                    # Special Handling for Column D in `adt_composite` & `idt`
                    if col == "D" and sheet_name in ["adt_composite", "idt"]:
                        if isinstance(value, (int, float)):  # If numerical, duplication is OK (log as Info)
                            if value in seen:
                                errors["Info"].append(f"[{sheet_name}] Duplicate numerical value at {cell_ref}: {value}")
                        else:  # If alphanumeric, duplication is NOT OK (log as Critical Error)
                            if value in seen:
                                errors["Critical"].append(f"[{sheet_name}] Duplicate non-numeric value at {cell_ref}: {value}")
                    else:
                        # General duplicate check for all other columns (log as Critical Error)
                        if value in seen:
                            errors["Critical"].append(f"[{sheet_name}] Duplicate value at {cell_ref}: {value}")
                    seen.add(value)  # Add value to seen set



        ### 🔵 Reference Validation ('excel_rule_4')###
        references = {
            ("ib_data", "F"): ("swc_info", "H"),
            ("ib_data", "D"): ("adt_primitive", "B"),
            ("ports", "I"): ("swc_info", "H"),
            ("ports", "H"): ("adt_primitive", "B")
        }
        for (sheet_name, col), (ref_sheet, ref_col) in references.items():
            if sheet_name not in wb.sheetnames or ref_sheet not in wb.sheetnames:
                continue
            sheet = wb[sheet_name]
            ref_values = {row[0] for row in wb[ref_sheet].iter_rows(min_row=2, min_col=ord(ref_col)-64, max_col=ord(ref_col)-64, values_only=True)}
            for row_idx, row in enumerate(sheet.iter_rows(
                    min_row=2, min_col=ord(col)-64, max_col=ord(col)-64, values_only=True), start=2):
                value = row[0]
                cell_ref = f"{col}{row_idx}"
                if value not in ref_values:
                    errors["Critical"].append(f"[{sheet_name}] Invalid reference at {cell_ref}: {value} (not in {ref_sheet}.{ref_col})")
    except Exception as e:
        errors["Critical"].append(f"Error reading Excel file: {str(e)}")
    return errors

    """
        Pre-defined rules for enum_list

    """
    enum_dictionary = {
        
    }


def print_colored_errors(errors):

   """ Prints errors in color-coded format. """

   for severity, msgs in errors.items():

       if msgs:

           color = RED if severity == "Critical" else (YELLOW if severity == "Warning" else BLUE)

           for msg in msgs:

               print(f"{color}{msg}{RESET}")  # Apply color and reset after message

def log_errors(errors, attempt_number):
   """ Logs validation errors with severity levels and timestamps. """
   timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
   with open("validation_log.txt", "a", encoding="utf-8") as log_file:
       log_file.write(f"\n=== Validation Attempt {attempt_number} at {timestamp} ===\n")
       for severity, msgs in errors.items():
           if msgs:
               log_file.write(f"\n[{severity} ERRORS]\n")
               for msg in msgs:
                   log_file.write(f"{msg}\n")
   print(f"{YELLOW}Errors logged in 'validation_log.txt'. Please fix them before retrying.{RESET}")

def generate_summary(initial_errors, final_errors, attempts):

   """ Generates a validation summary report. """

   fixed_errors = len(initial_errors["Critical"]) - len(final_errors["Critical"])

   summary_report = f"""

   ============================

   Validation Summary Report

   ============================

   📅 Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

   🔄 Validation Attempts: {attempts}

   ❌ Initial Critical Errors: {len(initial_errors["Critical"])}

   ⚠️ Initial Warnings: {len(initial_errors["Warning"])}

   ✅ Critical Errors Fixed: {fixed_errors}

   """

   print(summary_report)

   with open("validation_summary.txt", "w", encoding="utf-8") as summary_file:

       summary_file.write(summary_report)

   print(f"{GREEN}✔ Summary saved in 'validation_summary.txt'.{RESET}") 

def generate_html_report(errors, attempts):
    """
    Generates an HTML report for validation errors.
    """
    report_filename = f"validation_report_attempt_{attempts}.html"
    html_content = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel Validation Report - Attempt {attempts}</title>
    <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            h1, h2 {{ color: #333; }}
            .critical {{ color: red; font-weight: bold; }}
            .warning {{ color: orange; font-weight: bold; }}
            .info {{ color: blue; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
            th {{ background-color: #f2f2f2; }}
    </style>
    </head>
    <body>
    <h1>Excel Validation Report - Attempt {attempts}</h1>
    """
    # Add Critical Errors Section
    if errors["Critical"]:
        html_content += "<h2 class='critical'>❌ Critical Errors</h2><ul>"
        for error in errors["Critical"]:
            html_content += f"<li class='critical'>{error}</li>"
        html_content += "</ul>"
    # Add Warnings Section
    if errors["Warning"]:
        html_content += "<h2 class='warning'>⚠️ Warnings</h2><ul>"
        for error in errors["Warning"]:
            html_content += f"<li class='warning'>{error}</li>"
        html_content += "</ul>"
    # Add Info Section
    if errors["Info"]:
        html_content += "<h2 class='info'>ℹ️ Info Messages</h2><ul>"
        for error in errors["Info"]:
            html_content += f"<li class='info'>{error}</li>"
        html_content += "</ul>"
    # If no errors exist
    if not any(errors.values()):
        html_content += "<h2>✅ No validation errors found.</h2>"
    html_content += "</body></html>"
    # Write the HTML content to a file
    with open(report_filename, "w", encoding="utf-8") as file:
        file.write(html_content)
    print(f"\n📄 HTML Report Generated: {report_filename}")   